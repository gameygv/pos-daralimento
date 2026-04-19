import openpyxl, subprocess, json

wb = openpyxl.load_workbook('inventarios2026.xlsx', data_only=True)

def run_sql(sql):
    r = subprocess.run(['ssh', 'easypanel-vps',
        f'docker exec main_supabase-daralimento-db-1 psql -U supabase_admin -d postgres -t -A -c "{sql}"'],
        capture_output=True, text=True, timeout=15)
    return r.stdout.strip()

def parse_num(v):
    if not v: return 0
    s = str(v).replace('$', '').replace(' ', '')
    if ',' in s and '.' not in s:
        s = s.replace(',', '.')
    elif ',' in s and '.' in s:
        s = s.replace(',', '')
    try: return float(s)
    except: return 0

# Normalize client names
CLIENT_MAP = {
    'FAM HOCH': 'Familia Hoch',
    'FAM, HOCH': 'Familia Hoch',
    'FAM. HOCH': 'Familia Hoch',
    'FAM.  HOCH': 'Familia Hoch',
    'LA  PIZCA': 'LA PIZCA',
    'LORE ALVARADO': 'LORE ALVARADO',
    'PAG. WEB': 'PAG WEB',
}

# ============ CUENTAS POR COBRAR ============
ws = wb['Cuentas por cobrar']
cxc_by_client = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    fecha, cant, producto, peso, pub, prov, cliente, entregado = row[:8]
    if not producto or str(producto).strip() == 'PRODUCTO':
        continue

    cliente_name = str(cliente).strip() if cliente else 'Mostrador'
    cliente_name = CLIENT_MAP.get(cliente_name, cliente_name)

    pub_val = parse_num(pub)
    prov_val = parse_num(prov)
    price = pub_val if pub_val > 0 else prov_val

    if cliente_name not in cxc_by_client:
        cxc_by_client[cliente_name] = {'total': 0, 'items': 0, 'fecha': str(fecha)[:10] if fecha else '2026-04-01'}
    cxc_by_client[cliente_name]['total'] += price
    cxc_by_client[cliente_name]['items'] += 1
    # Use latest date
    if fecha and str(fecha)[:10] > cxc_by_client[cliente_name]['fecha']:
        cxc_by_client[cliente_name]['fecha'] = str(fecha)[:10]

print("=== CUENTAS POR COBRAR (agrupadas por cliente) ===")
sql = []
folio_counter = 9000  # Start at high number to avoid conflicts

for cliente, info in sorted(cxc_by_client.items()):
    total = round(info['total'], 2)
    if total <= 0:
        continue
    folio_counter += 1
    fecha = info['fecha'] if info['fecha'] != 'None' else '2026-04-01'
    safe_cliente = cliente.replace("'", "''")
    print(f"  {cliente}: {info['items']} items, total=${total}, fecha={fecha}")

    sql.append(f"""INSERT INTO notas (folio, folio_display, fecha, hora, cliente, vendedor, total, pagado, metodo_pago, pago_status, entrega_status, notas_pago)
VALUES ({folio_counter}, 'IMP-{folio_counter}', '{fecha}', '00:00:00', '{safe_cliente}', 'Importacion', {total}, 0, 'efectivo', 'pendiente', 'entregado', 'Importado desde Excel - Cuentas por cobrar')
ON CONFLICT DO NOTHING;""")

# ============ ORDENES PAGADAS ============
ws2 = wb['Ordenes de compra']
orders_by_client = {}
for row in ws2.iter_rows(min_row=4, values_only=True):
    fecha, num, producto, peso, pub, prov, cliente, entregado, pagado = row[:9]
    if not producto or str(producto).strip() == 'PRODUCTO':
        continue

    is_paid = pagado and str(pagado).strip().upper() == 'X'
    is_delivered = entregado and str(entregado).strip().upper() == 'X'

    cliente_name = str(cliente).strip() if cliente else 'Mostrador'
    cliente_name = CLIENT_MAP.get(cliente_name, cliente_name)

    pub_val = parse_num(pub)
    prov_val = parse_num(prov)
    price = pub_val if pub_val > 0 else prov_val

    key = (cliente_name, 'paid' if is_paid else 'unpaid')
    if key not in orders_by_client:
        orders_by_client[key] = {'total': 0, 'items': 0, 'fecha': str(fecha)[:10] if fecha else '2026-04-01', 'delivered': is_delivered}
    orders_by_client[key]['total'] += price
    orders_by_client[key]['items'] += 1

print("\n=== ORDENES DE COMPRA ===")
for (cliente, status), info in sorted(orders_by_client.items()):
    total = round(info['total'], 2)
    if total <= 0:
        continue
    folio_counter += 1
    fecha = info['fecha'] if info['fecha'] != 'None' else '2026-04-01'
    safe_cliente = cliente.replace("'", "''")
    pago_status = 'pagado' if status == 'paid' else 'pendiente'
    pagado_val = total if status == 'paid' else 0
    entrega = 'entregado'
    print(f"  {cliente} ({status}): {info['items']} items, total=${total}")

    sql.append(f"""INSERT INTO notas (folio, folio_display, fecha, hora, cliente, vendedor, total, pagado, metodo_pago, pago_status, entrega_status, notas_pago)
VALUES ({folio_counter}, 'IMP-{folio_counter}', '{fecha}', '00:00:00', '{safe_cliente}', 'Importacion', {total}, {pagado_val}, 'efectivo', '{pago_status}', '{entrega}', 'Importado desde Excel - Ordenes de compra')
ON CONFLICT DO NOTHING;""")

with open('import_cxc.sql', 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql))
print(f'\nSQL: {len(sql)} statements written to import_cxc.sql')
